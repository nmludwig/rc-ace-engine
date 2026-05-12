#!/usr/bin/env python3
"""
RingCentral ACE Engine — Web Server
Full pipeline: OAuth → transcript download → ACE analysis → HTML microsite
"""

import os, re, time, threading, uuid, secrets, json
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode

from flask import (Flask, render_template, request, jsonify,
                   send_file, session, redirect, url_for, abort)

try:
    from flask_cors import CORS
except ImportError:
    CORS = None

import requests

# ── Redis job store ────────────────────────────────────────────
try:
    import redis
    REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379")
    _redis = redis.from_url(REDIS_URL, decode_responses=True)
    _redis.ping()
    USE_REDIS = True
except Exception:
    USE_REDIS = False
    _jobs_mem = {}   # fallback in-memory store

JOB_TTL = 60 * 60 * 24  # 24 hours

def job_set(job_id, data):
    if USE_REDIS:
        _redis.setex(f"job:{job_id}", JOB_TTL, json.dumps(data))
    else:
        _jobs_mem[job_id] = data

def job_get(job_id):
    if USE_REDIS:
        raw = _redis.get(f"job:{job_id}")
        return json.loads(raw) if raw else None
    return _jobs_mem.get(job_id)

def job_update(job_id, patch):
    data = job_get(job_id) or {}
    data.update(patch)
    job_set(job_id, data)

def job_append_log(job_id, msg, level="info"):
    data = job_get(job_id) or {}
    log = data.get("log", [])
    log.append({"t": datetime.now().strftime("%H:%M:%S"), "msg": msg, "level": level})
    data["log"] = log[-200:]   # keep last 200 entries
    job_set(job_id, data)

# ── RingCentral config ─────────────────────────────────────────
RC_CLIENT_ID     = os.environ.get("RC_CLIENT_ID", "")
RC_CLIENT_SECRET = os.environ.get("RC_CLIENT_SECRET", "")
RC_REDIRECT_URI  = os.environ.get("RC_REDIRECT_URI",
                   "https://rc-ace-engine.onrender.com/oauth/callback")
RC_AUTH_URL  = "https://platform.ringcentral.com/restapi/oauth/authorize"
RC_TOKEN_URL = "https://platform.ringcentral.com/restapi/oauth/token"

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# ── Paths ──────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
OUTPUT_DIR  = BASE_DIR / "outputs"
MICROSITE_DIR = BASE_DIR / "microsites"
OUTPUT_DIR.mkdir(exist_ok=True)
MICROSITE_DIR.mkdir(exist_ok=True)

# ── Flask ──────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", secrets.token_hex(32))
if CORS:
    CORS(app)

# ══════════════════════════════════════════════════════════════
# ROUTES — Auth
# ══════════════════════════════════════════════════════════════

@app.route("/")
def index():
    authed = bool(session.get("rc_token"))
    return render_template("index.html",
                           authed=authed,
                           display_name=session.get("rc_display_name", ""))

@app.route("/login")
def login():
    state = secrets.token_urlsafe(16)
    session["oauth_state"] = state
    params = {
        "response_type": "code",
        "client_id": RC_CLIENT_ID,
        "redirect_uri": RC_REDIRECT_URI,
        "state": state,
        "scope": "ReadCallLog ReadCallRecording RingSense ReadAccounts Analytics ReadContacts"
    }
    return redirect(RC_AUTH_URL + "?" + urlencode(params))

@app.route("/oauth/callback")
def oauth_callback():
    error = request.args.get("error")
    if error:
        return render_template("error.html", message=f"Login failed: {error}"), 400
    code  = request.args.get("code")
    state = request.args.get("state")
    if not code:
        return render_template("error.html", message="No auth code received."), 400
    if state != session.get("oauth_state"):
        return render_template("error.html", message="Invalid state parameter."), 400
    try:
        resp = requests.post(RC_TOKEN_URL,
                             auth=(RC_CLIENT_ID, RC_CLIENT_SECRET),
                             data={"grant_type": "authorization_code",
                                   "code": code, "redirect_uri": RC_REDIRECT_URI},
                             timeout=15)
        resp.raise_for_status()
    except requests.exceptions.HTTPError as e:
        code_n = e.response.status_code if e.response else "?"
        return render_template("error.html", message=f"Token exchange failed (HTTP {code_n})."), 400
    except Exception as e:
        return render_template("error.html", message=f"Connection error: {e}"), 503

    data  = resp.json()
    token = data["access_token"]
    display_name = ""
    account_id   = "~"
    try:
        me = requests.get(
            "https://platform.ringcentral.com/restapi/v1.0/account/~/extension/~",
            headers={"Authorization": "Bearer " + token}, timeout=15).json()
        display_name = me.get("name", "") or me.get("contact", {}).get("firstName", "")
        acct = requests.get(
            "https://platform.ringcentral.com/restapi/v1.0/account/~/call-log",
            headers={"Authorization": "Bearer " + token},
            params={"perPage": 1}, timeout=15)
        m = re.search(r"account/(\d+)", acct.json().get("uri", ""))
        if m:
            account_id = m.group(1)
    except Exception:
        pass

    session["rc_token"]         = token
    session["rc_account_id"]    = account_id
    session["rc_display_name"]  = display_name
    session["rc_refresh_token"] = data.get("refresh_token", "")
    session["rc_token_time"]    = datetime.now().timestamp()
    session.pop("oauth_state", None)
    return redirect(url_for("index"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))

# ══════════════════════════════════════════════════════════════
# ROUTES — Download job (transcript export)
# ══════════════════════════════════════════════════════════════

@app.route("/api/start-download", methods=["POST"])
def api_start_download():
    if not session.get("rc_token"):
        return jsonify({"error": "Not authenticated."}), 401
    data = request.json or {}
    customer_name = (data.get("customer_name") or "Customer").strip()
    date_from     = data.get("date_from", "")
    date_to       = data.get("date_to", "")
    if not date_from or not date_to:
        return jsonify({"error": "date_from and date_to are required."}), 400

    job_id = str(uuid.uuid4())
    job_set(job_id, {
        "type": "download", "status": "running", "progress": 0,
        "log": [], "files": {}, "error": None, "summary": None,
        "customer_name": customer_name,
        "transcript_text": ""
    })
    threading.Thread(
        target=run_download_job,
        args=(job_id, session["rc_token"], session["rc_account_id"],
              session.get("rc_refresh_token", ""),
              customer_name, date_from, date_to),
        daemon=True).start()
    return jsonify({"job_id": job_id})

@app.route("/api/job/<job_id>")
def api_job_status(job_id):
    job = job_get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "type":     job.get("type", "download"),
        "status":   job["status"],
        "progress": job["progress"],
        "log":      job.get("log", [])[-50:],
        "files":    list(job.get("files", {}).keys()),
        "error":    job.get("error"),
        "summary":  job.get("summary"),
        "microsite_url": job.get("microsite_url"),
    })

@app.route("/api/download/<job_id>/<file_type>")
def api_download(job_id, file_type):
    job = job_get(job_id)
    if not job or file_type not in job.get("files", {}):
        return jsonify({"error": "File not found"}), 404
    path = Path(job["files"][file_type])
    if not path.exists():
        return jsonify({"error": "File no longer available"}), 404
    return send_file(str(path), as_attachment=True, download_name=path.name)

# ══════════════════════════════════════════════════════════════
# ROUTES — ACE analysis job
# ══════════════════════════════════════════════════════════════

@app.route("/api/start-ace", methods=["POST"])
def api_start_ace():
    if not session.get("rc_token"):
        return jsonify({"error": "Not authenticated."}), 401
    data          = request.json or {}
    source_job_id = data.get("source_job_id", "")
    customer_name = (data.get("customer_name") or "Customer").strip()
    industry      = (data.get("industry") or "Equipment").strip()

    source_job = job_get(source_job_id) if source_job_id else None
    transcript_text = ""
    if source_job:
        transcript_text = source_job.get("transcript_text", "")

    if not transcript_text:
        return jsonify({"error": "No transcript text found. Please run a download first."}), 400
    if not ANTHROPIC_API_KEY:
        return jsonify({"error": "ANTHROPIC_API_KEY not configured on server."}), 500

    ace_job_id = str(uuid.uuid4())
    job_set(ace_job_id, {
        "type": "ace", "status": "running", "progress": 0,
        "log": [], "files": {}, "error": None, "summary": None,
        "customer_name": customer_name, "industry": industry,
        "microsite_url": None
    })
    threading.Thread(
        target=run_ace_job,
        args=(ace_job_id, transcript_text, customer_name, industry),
        daemon=True).start()
    return jsonify({"job_id": ace_job_id})

@app.route("/ace/<slug>")
def serve_microsite(slug):
    """Serve a generated ACE microsite."""
    path = MICROSITE_DIR / slug / "index.html"
    if not path.exists():
        abort(404)
    return path.read_text(encoding="utf-8")

@app.route("/ace/<slug>/download/<file_type>")
def download_ace_file(slug, file_type):
    """Download PDF, scorecard, or keywords from a microsite run."""
    allowed = {"pdf": ".pdf", "scorecard": "_Scorecard.txt", "keywords": "_Keywords.txt"}
    if file_type not in allowed:
        abort(404)
    folder = MICROSITE_DIR / slug
    matches = list(folder.glob(f"*{allowed[file_type]}"))
    if not matches:
        abort(404)
    return send_file(str(matches[0]), as_attachment=True, download_name=matches[0].name)

# ══════════════════════════════════════════════════════════════
# DOWNLOAD JOB WORKER
# ══════════════════════════════════════════════════════════════

def refresh_rc_token(refresh_token):
    try:
        resp = requests.post(RC_TOKEN_URL,
                             auth=(RC_CLIENT_ID, RC_CLIENT_SECRET),
                             data={"grant_type": "refresh_token",
                                   "refresh_token": refresh_token},
                             timeout=15)
        if resp.status_code == 200:
            d = resp.json()
            return d.get("access_token"), d.get("refresh_token", refresh_token)
    except Exception:
        pass
    return None, refresh_token

def run_download_job(job_id, token, account_id, refresh_token,
                     customer_name, date_from, date_to):
    try:
        job_append_log(job_id, f"Starting download for {customer_name}")
        job_append_log(job_id, f"Date range: {date_from} → {date_to}")
        job_update(job_id, {"progress": 5})

        records = []
        page = 1
        while True:
            resp = requests.get(
                f"https://platform.ringcentral.com/restapi/v1.0/account/{account_id}/call-log",
                headers={"Authorization": "Bearer " + token},
                params={"view": "Detailed",
                        "dateFrom": date_from + "T00:00:00Z",
                        "dateTo":   date_to   + "T23:59:59Z",
                        "type": "Voice", "withRecording": "true",
                        "perPage": 250, "page": page},
                timeout=30)
            resp.raise_for_status()
            data = resp.json()
            records.extend(data.get("records", []))
            if not data.get("navigation", {}).get("nextPage"):
                break
            page += 1
            time.sleep(0.25)

        job_update(job_id, {"progress": 15})
        job_append_log(job_id, f"Found {len(records)} recorded calls", "ok")

        if not records:
            job_update(job_id, {"status": "done", "progress": 100,
                                "summary": {"total": 0, "transcripts": 0}})
            job_append_log(job_id, "No recorded calls found in this date range.", "warn")
            return

        total = len(records)
        job_append_log(job_id, f"Fetching RingSense transcripts for {total} calls…")
        job_append_log(job_id, f"~{round(total * 1.5 / 60, 1)} minutes estimated. Please wait…", "warn")

        transcript_records = []
        transcript_parts   = []   # for ACE analysis
        with_transcripts   = 0

        for i, call in enumerate(records):
            recording_id = call.get("recording", {}).get("id")
            if not recording_id:
                continue

            job_update(job_id, {"progress": 15 + int(70 * (i / max(total, 1)))})

            insights = None
            url = (f"https://platform.ringcentral.com/ai/ringsense/v1/public"
                   f"/accounts/{account_id}/domains/pbx/records/{recording_id}/insights")
            while True:
                try:
                    r = requests.get(url, headers={"Authorization": "Bearer " + token},
                                     timeout=30)
                    if r.status_code == 429:
                        job_append_log(job_id, "Rate limit — waiting 65s…", "warn")
                        time.sleep(65)
                        continue
                    if r.status_code == 200:
                        insights = r.json()
                    break
                except Exception:
                    break

            if insights:
                with_transcripts += 1

            speaker_map = {}
            if insights:
                for sp in insights.get("speakerInfo", []):
                    sid  = sp.get("speakerId", "")
                    name = sp.get("name", "") or sp.get("phoneNumber", sid)
                    if sid and name:
                        speaker_map[sid] = name

            utterances = (insights or {}).get("insights", {}).get("Transcript", [])
            lines = []
            for u in utterances:
                sid  = u.get("speakerId", "?")
                name = speaker_map.get(sid, sid)
                txt  = u.get("text", "").strip()
                s    = u.get("start", 0)
                mm   = str(int(s // 60)).zfill(2)
                ss   = str(int(s % 60)).zfill(2)
                lines.append(f"[{mm}:{ss}] {name}: {txt}")

            summary_list   = (insights or {}).get("insights", {}).get("Summary", [])
            sentiment_list = (insights or {}).get("insights", {}).get("Sentiment", [])
            summary   = summary_list[0].get("value", "")   if summary_list   else ""
            sentiment = sentiment_list[0].get("value", "") if sentiment_list else ""

            rec = call.get("recording", {})
            row = {
                "call_id":      call.get("id", ""),
                "start_time":   call.get("startTime", ""),
                "duration_sec": call.get("duration", 0),
                "direction":    call.get("direction", ""),
                "from_number":  call.get("from", {}).get("phoneNumber", ""),
                "from_name":    call.get("from", {}).get("name", ""),
                "to_number":    call.get("to", {}).get("phoneNumber", ""),
                "to_name":      call.get("to", {}).get("name", ""),
                "recording_id": rec.get("id", ""),
                "has_transcript": insights is not None,
                "sentiment":    sentiment,
                "summary":      summary,
                "transcript":   "\n".join(lines),
            }
            transcript_records.append(row)

            # Build combined text for ACE analysis
            if insights:
                try:
                    dt = datetime.fromisoformat(row["start_time"].replace("Z", "+00:00"))
                    date_str = dt.strftime("%B %d, %Y")
                except Exception:
                    date_str = row["start_time"][:10]
                header = (f"\nCALL {i+1} OF {total}\n"
                          f"Inbound call from {row['from_name'] or row['from_number']}\n"
                          f"{date_str} | Duration: {row['duration_sec']}s\n"
                          f"Inbound\n{sentiment}\n\nAI SUMMARY\n{summary}\n\n"
                          f"FULL TRANSCRIPT\n" + "\n".join(lines))
                transcript_parts.append(header)

            if (i + 1) % 10 == 0:
                job_append_log(job_id,
                    f"Processed {i+1}/{total} calls ({with_transcripts} transcripts so far)")

            if refresh_token and (i + 1) % 50 == 0:
                new_token, refresh_token = refresh_rc_token(refresh_token)
                if new_token:
                    token = new_token
                    job_append_log(job_id, "Token refreshed.", "ok")

            time.sleep(1.5)

        job_append_log(job_id, f"{with_transcripts} of {total} calls have transcripts", "ok")
        job_update(job_id, {"progress": 88})

        slug       = re.sub(r"[^a-z0-9]+", "_", customer_name.lower()).strip("_")
        date_stamp = datetime.now().strftime("%Y%m%d")

        # Build Excel
        job_append_log(job_id, "Building Excel spreadsheet…")
        xlsx_path = OUTPUT_DIR / f"transcripts_{slug}_{date_stamp}.xlsx"
        build_excel(transcript_records, customer_name, date_from, date_to, xlsx_path)
        job_append_log(job_id, f"Excel ready.", "ok")
        job_update(job_id, {"progress": 93})

        # Build PDF
        job_append_log(job_id, "Building PDF…")
        pdf_path = OUTPUT_DIR / f"transcripts_{slug}_{date_stamp}.pdf"
        build_pdf(transcript_records, customer_name, date_from, date_to, pdf_path)
        job_append_log(job_id, "PDF ready.", "ok")

        # Save combined transcript text for ACE
        full_transcript = "\n".join(transcript_parts)

        job_update(job_id, {
            "progress": 100,
            "status":   "done",
            "files":    {"xlsx": str(xlsx_path), "pdf": str(pdf_path)},
            "transcript_text": full_transcript[:500000],  # cap at 500k chars
            "summary": {
                "total":        len(records),
                "transcripts":  with_transcripts,
                "customer":     customer_name,
                "date_from":    date_from,
                "date_to":      date_to,
            }
        })
        job_append_log(job_id, "Download complete — files ready.", "ok")
        job_append_log(job_id, "You can now run the ACE Analysis.", "ok")

    except Exception as e:
        job_update(job_id, {"status": "error", "error": str(e)})
        job_append_log(job_id, f"Error: {e}", "error")

# ══════════════════════════════════════════════════════════════
# ACE ANALYSIS JOB WORKER
# ══════════════════════════════════════════════════════════════

def run_ace_job(job_id, transcript_text, customer_name, industry):
    try:
        from ace_engine import run_ace_analysis, build_html, build_scorecard, build_keywords
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.colors import HexColor, white

        job_append_log(job_id, f"Starting ACE analysis for {customer_name}…")
        job_update(job_id, {"progress": 5})

        slug = re.sub(r"[^a-z0-9]+", "_", customer_name.lower()).strip("_")

        # Count calls and extract date range
        call_count = max(len(re.findall(r'CALL \d+ OF', transcript_text)), 1)
        dates = re.findall(
            r'(January|February|March|April|May|June|July|August|'
            r'September|October|November|December)\s+\d{1,2},\s+\d{4}',
            transcript_text)
        date_range = (f"{dates[0]} – {dates[-1]}" if len(dates) >= 2
                      else (dates[0] if dates else datetime.now().strftime("%B %Y")))

        job_append_log(job_id, f"Detected ~{call_count} calls · {date_range}")
        job_update(job_id, {"progress": 10})

        # Run Claude analysis
        job_append_log(job_id, "Sending transcripts to Claude for analysis (~60s)…")
        data = run_ace_analysis(transcript_text, customer_name, industry,
                                call_count, date_range)
        job_append_log(job_id, "Analysis complete.", "ok")
        job_update(job_id, {"progress": 55})

        # Create output folder under microsites/
        run_ts    = datetime.now().strftime("%Y%m%d_%H%M")
        site_slug = f"{slug}_{run_ts}"
        site_dir  = MICROSITE_DIR / site_slug
        site_dir.mkdir(parents=True, exist_ok=True)

        # Save JSON
        with open(str(site_dir / "analysis.json"), "w") as f:
            json.dump(data, f, indent=2)

        # HTML microsite
        job_append_log(job_id, "Building HTML microsite…")
        build_html(data, site_dir, slug)
        job_append_log(job_id, "Microsite ready.", "ok")
        job_update(job_id, {"progress": 70})

        # Leave-behind PDF
        job_append_log(job_id, "Building PDF leave-behind…")
        pdf_path = str(site_dir / f"ACE_{slug}_Leave_Behind.pdf")
        build_ace_pdf(data, pdf_path)
        job_append_log(job_id, "PDF ready.", "ok")
        job_update(job_id, {"progress": 82})

        # Scorecard
        job_append_log(job_id, "Writing scorecard…")
        sc_path = str(site_dir / f"ACE_{slug}_Scorecard.txt")
        build_scorecard(data, sc_path, customer_name)
        job_append_log(job_id, "Scorecard ready.", "ok")
        job_update(job_id, {"progress": 90})

        # Keywords
        job_append_log(job_id, "Writing keywords…")
        kw_path = str(site_dir / f"ACE_{slug}_Keywords.txt")
        build_keywords(data, kw_path, customer_name)
        job_append_log(job_id, "Keywords ready.", "ok")
        job_update(job_id, {"progress": 98})

        microsite_url = f"/ace/{site_slug}"

        job_update(job_id, {
            "status":   "done",
            "progress": 100,
            "microsite_url": microsite_url,
            "files": {
                "pdf":       pdf_path,
                "scorecard": sc_path,
                "keywords":  kw_path,
            },
            "summary": {
                "customer": customer_name,
                "industry": industry,
                "site_slug": site_slug,
            }
        })
        job_append_log(job_id, f"All done! Microsite live at {microsite_url}", "ok")

    except Exception as e:
        import traceback
        job_update(job_id, {"status": "error", "error": str(e)})
        job_append_log(job_id, f"Error: {e}", "error")
        job_append_log(job_id, traceback.format_exc(), "error")


def build_ace_pdf(data, out_path):
    """Thin wrapper — imports reportlab and calls the PDF builder from ace_engine."""
    from ace_engine import build_leave_behind_pdf
    build_leave_behind_pdf(data, out_path)


# ══════════════════════════════════════════════════════════════
# EXCEL + PDF BUILDERS (transcript exports — unchanged from v1)
# ══════════════════════════════════════════════════════════════

def build_excel(records, customer_name, date_from, date_to, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    RC_ORANGE = "FF6A00"; DARK = "1A1A1A"; WHITE = "FFFFFF"; LIGHT_GREY = "F5F5F5"
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Summary"
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"RingCentral ACE Transcript Report — {customer_name}"
    c.font  = Font(name="Calibri", size=16, bold=True, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=RC_ORANGE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = f"{date_from} to {date_to} | Generated {datetime.now().strftime('%B %d, %Y')} | Confidential"
    c2.font  = Font(name="Calibri", size=10, color="555555")
    c2.fill  = PatternFill("solid", fgColor=LIGHT_GREY)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    with_trans = len([r for r in records if r["has_transcript"]])
    total_sec  = sum(r["duration_sec"] for r in records)
    hrs, rem   = divmod(total_sec, 3600); mins = rem // 60

    ws.append([])
    for row_idx, (label, value) in enumerate([
        ("Total Calls", str(len(records))), ("With Transcripts", str(with_trans)),
        ("Without Transcripts", str(len(records) - with_trans)),
        ("Total Talk Time", f"{hrs}h {mins}m"),
        ("Date From", date_from), ("Date To", date_to)
    ], start=4):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=1).font  = Font(name="Calibri", size=11, bold=True, color=DARK)
        ws.cell(row=row_idx, column=1).fill  = PatternFill("solid", fgColor=LIGHT_GREY)
        ws.cell(row=row_idx, column=2).value = value
        ws.cell(row=row_idx, column=2).font  = Font(name="Calibri", size=11, color=DARK)
    ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 20

    ws2 = wb.create_sheet("All Calls")
    headers = ["Date","Time","Direction","Duration","From Name","From Number",
               "To Name","To Number","Has Transcript","Sentiment","AI Summary"]
    widths  = [14,10,12,10,22,18,22,18,14,12,60]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = h
        cell.font  = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=RC_ORANGE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 24

    for ri, rec in enumerate(records, 2):
        try:
            dt = datetime.fromisoformat(rec["start_time"].replace("Z", "+00:00"))
            ds = dt.strftime("%Y-%m-%d"); ts = dt.strftime("%I:%M %p")
        except Exception:
            ds = rec["start_time"][:10]; ts = ""
        dm, ds2 = divmod(int(rec["duration_sec"]), 60)
        for col, val in enumerate([ds, ts, rec["direction"], f"{dm}m {ds2}s",
                rec["from_name"], rec["from_number"], rec["to_name"], rec["to_number"],
                "Yes" if rec["has_transcript"] else "No",
                rec["sentiment"], rec["summary"]], 1):
            cell = ws2.cell(row=ri, column=col)
            cell.value = val
            cell.font  = Font(name="Calibri", size=10, color=DARK)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 11))
            cell.border = border
        ws2.row_dimensions[ri].height = 15

    ws3 = wb.create_sheet("Transcripts")
    th = ["Date","From Name","To Name","Duration","Sentiment","AI Summary","Full Transcript"]
    tw = [14,22,22,10,12,60,100]
    for col, (h, w) in enumerate(zip(th, tw), 1):
        cell = ws3.cell(row=1, column=col)
        cell.value = h
        cell.font  = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=RC_ORANGE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws3.column_dimensions[get_column_letter(col)].width = w

    for ri, rec in enumerate([r for r in records if r["has_transcript"]], 2):
        try:
            dt = datetime.fromisoformat(rec["start_time"].replace("Z", "+00:00"))
            ds = dt.strftime("%Y-%m-%d")
        except Exception:
            ds = rec["start_time"][:10]
        dm, ds2 = divmod(int(rec["duration_sec"]), 60)
        for col, val in enumerate([ds, rec["from_name"], rec["to_name"],
                f"{dm}m {ds2}s", rec["sentiment"], rec["summary"], rec["transcript"]], 1):
            cell = ws3.cell(row=ri, column=col)
            cell.value = val
            cell.font  = Font(name="Calibri", size=10, color=DARK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws3.row_dimensions[ri].height = 60

    wb.save(out_path)


def build_pdf(records, customer_name, date_from, date_to, out_path):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable, KeepTogether)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfgen import canvas as rc_canvas

    RC_ORANGE = HexColor("#FF6A00"); DARK = HexColor("#1A1A1A"); GREY = HexColor("#666666")
    BG_GREEN  = HexColor("#E8F8EF"); BG_RED   = HexColor("#FEECEC")
    BG_GREY   = HexColor("#F5F5F5"); BG_BLUE  = HexColor("#E6F1FB")
    BG_ORANGE = HexColor("#FFF3E8"); RULE     = HexColor("#E0E0E0")
    TX_GREEN  = HexColor("#1B6B35"); TX_RED   = HexColor("#8B1A1A")
    TX_GREY   = HexColor("#444444"); TX_BLUE  = HexColor("#0B4F8A")

    def ps(name, **kw): return ParagraphStyle(name, **kw)

    class NumCanvas(rc_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs); self._saved_page_states = []
        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__)); self._startPage()
        def save(self):
            n = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state); self._draw_page_number(n); super().showPage()
            super().save()
        def _draw_page_number(self, n):
            self.setFont("Helvetica", 8); self.setFillColor(HexColor("#AAAAAA"))
            self.drawRightString(letter[0]-0.6*inch, 0.4*inch,
                                 f"Page {self._pageNumber} of {n}")

    doc   = SimpleDocTemplate(str(out_path), pagesize=letter,
                              leftMargin=0.75*inch, rightMargin=0.75*inch,
                              topMargin=0.75*inch, bottomMargin=0.75*inch)
    story = []

    cover_data = [[
        Paragraph("<b>RingCentral ACE Transcript Report</b>",
                  ps("H1", fontName="Helvetica-Bold", fontSize=18,
                     textColor=HexColor("#FFFFFF"), leading=22)),
        Paragraph(customer_name,
                  ps("H2", fontName="Helvetica-Bold", fontSize=14,
                     textColor=HexColor("#FFD0A8"), leading=18, alignment=TA_RIGHT)),
    ]]
    ct = Table(cover_data, colWidths=[doc.width*0.6, doc.width*0.4])
    ct.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), RC_ORANGE),
        ("TOPPADDING",(0,0),(-1,-1), 14), ("BOTTOMPADDING",(0,0),(-1,-1), 14),
        ("LEFTPADDING",(0,0),(0,0), 18),  ("RIGHTPADDING",(-1,0),(-1,-1), 18),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story.append(ct); story.append(Spacer(1,4))

    sub = Table([[Paragraph(
        f"{date_from} → {date_to} | Generated {datetime.now().strftime('%B %d, %Y')} | Confidential",
        ps("SU", fontName="Helvetica", fontSize=8.5, textColor=HexColor("#555555"),
           alignment=TA_CENTER))]],
        colWidths=[doc.width])
    sub.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),HexColor("#F5F5F5")),
                              ("TOPPADDING",(0,0),(-1,-1),7),
                              ("BOTTOMPADDING",(0,0),(-1,-1),7)]))
    story.append(sub); story.append(Spacer(1,16))

    call_num_s  = ps("CN", fontName="Helvetica", fontSize=8, textColor=GREY, leading=12)
    call_title_s= ps("CT", fontName="Helvetica-Bold", fontSize=13,
                     textColor=DARK, leading=16, spaceAfter=4)
    meta_s      = ps("ME", fontName="Helvetica", fontSize=9,
                     textColor=GREY, leading=12, spaceAfter=6)
    sum_lbl_s   = ps("SL", fontName="Helvetica-Bold", fontSize=8,
                     textColor=RC_ORANGE, leading=14, spaceBefore=10)
    sum_txt_s   = ps("ST", fontName="Helvetica", fontSize=10, textColor=DARK, leading=14)
    tr_lbl_s    = ps("TL", fontName="Helvetica-Bold", fontSize=8,
                     textColor=GREY, leading=14, spaceBefore=10)
    no_tr_s     = ps("NT", fontName="Helvetica-Oblique", fontSize=9,
                     textColor=GREY, leading=12)
    utt_s       = ps("UT", fontName="Helvetica", fontSize=9,
                     textColor=DARK, leading=13, leftIndent=10, spaceAfter=4)
    sp_styles   = [
        ps("SA", fontName="Helvetica-Bold", fontSize=9,
           textColor=HexColor("#0B4F8A"), leading=12, spaceBefore=6),
        ps("SB", fontName="Helvetica-Bold", fontSize=9,
           textColor=HexColor("#8B1A1A"), leading=12, spaceBefore=6),
        ps("SC", fontName="Helvetica-Bold", fontSize=9,
           textColor=HexColor("#1B6B35"), leading=12, spaceBefore=6),
        ps("SD", fontName="Helvetica-Bold", fontSize=9,
           textColor=HexColor("#5A189A"), leading=12, spaceBefore=6),
    ]

    calls_to_show = [r for r in records if r["has_transcript"]] or records
    for idx, row in enumerate(calls_to_show, 1):
        from_name = row["from_name"] or row["from_number"] or "Unknown"
        to_name   = row["to_name"]   or row["to_number"]   or "Unknown"
        try:
            dt = datetime.fromisoformat(row["start_time"].replace("Z","+00:00"))
            date_str = dt.strftime("%B %d, %Y"); time_str = dt.strftime("%I:%M %p")
        except Exception:
            date_str = row["start_time"][:10]; time_str = ""

        dur_m, dur_s = divmod(int(row["duration_sec"]), 60)
        sl = row["sentiment"].lower()
        if "positive" in sl: sent_bg,sent_fg,sent_lbl = BG_GREEN,TX_GREEN,"Positive"
        elif "negative" in sl: sent_bg,sent_fg,sent_lbl = BG_RED,TX_RED,"Negative"
        else: sent_bg,sent_fg,sent_lbl = BG_GREY,TX_GREY,"Neutral"

        dir_bg = BG_BLUE  if row["direction"]=="Inbound" else BG_ORANGE
        dir_fg = TX_BLUE  if row["direction"]=="Inbound" else HexColor("#9B4A00")

        block = []
        block.append(Paragraph(f"CALL {idx} OF {len(calls_to_show)}", call_num_s))
        block.append(Paragraph(
            f"Inbound call from <b>{from_name}</b>"
            if row["direction"]=="Inbound" else f"Outbound call to <b>{to_name}</b>",
            call_title_s))
        meta = [p for p in [date_str, time_str, f"Duration: {dur_m}m {dur_s}s",
                             f"From: {from_name}", f"To: {to_name}"] if p]
        block.append(Paragraph(" &nbsp;|&nbsp; ".join(meta), meta_s))

        bt = Table([[
            Paragraph(f"<b>{row['direction']}</b>",
                      ps("DB",fontSize=8.5,fontName="Helvetica-Bold",
                         textColor=dir_fg,alignment=TA_CENTER)),
            Paragraph(f"<b>{sent_lbl}</b>",
                      ps("SB2",fontSize=8.5,fontName="Helvetica-Bold",
                         textColor=sent_fg,alignment=TA_CENTER)),
            Paragraph("", ps("SP",fontSize=8)),
        ]], colWidths=[1.0*inch, 1.0*inch, doc.width-2.0*inch])
        bt.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(0,0),dir_bg), ("BACKGROUND",(1,0),(1,0),sent_bg),
            ("TOPPADDING",(0,0),(1,0),5), ("BOTTOMPADDING",(0,0),(1,0),5),
            ("LEFTPADDING",(0,0),(1,0),10), ("RIGHTPADDING",(0,0),(1,0),10),
        ]))
        block.append(Spacer(1,4)); block.append(bt)

        if row["summary"]:
            block.append(Paragraph("AI SUMMARY", sum_lbl_s))
            st2 = Table([[Paragraph(row["summary"], sum_txt_s)]], colWidths=[doc.width])
            st2.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), BG_ORANGE),
                ("TOPPADDING",(0,0),(-1,-1),10), ("BOTTOMPADDING",(0,0),(-1,-1),10),
                ("LEFTPADDING",(0,0),(-1,-1),14), ("RIGHTPADDING",(0,0),(-1,-1),14),
                ("LINEBEFORE",(0,0),(0,-1),3,RC_ORANGE),
            ]))
            block.append(st2)

        if row["transcript"]:
            block.append(Paragraph("FULL TRANSCRIPT", tr_lbl_s))
            unique_sp = list(dict.fromkeys(
                line.split(": ")[0].split("] ")[-1].strip()
                for line in row["transcript"].split("\n") if ": " in line))
            sp_map = {sp: sp_styles[i % len(sp_styles)] for i,sp in enumerate(unique_sp)}
            for line in row["transcript"].split("\n"):
                if not line.strip(): continue
                m = re.match(r"^(\[\d+:\d+\])\s+(.+?):\s+(.+)$", line)
                if m:
                    ts2, speaker, text = m.group(1), m.group(2), m.group(3)
                    safe = text.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                    block.append(Paragraph(
                        f"<b>{speaker}</b> <font size='7' color='#BBBBBB'>{ts2}</font>",
                        sp_map.get(speaker, sp_styles[0])))
                    block.append(Paragraph(safe, utt_s))
                else:
                    safe = line.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                    block.append(Paragraph(safe, utt_s))
        else:
            block.append(Spacer(1,6))
            block.append(Paragraph(
                "No transcript available — extension may not have a RingSense license.",
                no_tr_s))

        block.append(Spacer(1, 0.15*inch))
        block.append(HRFlowable(width="100%", thickness=0.5, color=RULE))
        block.append(Spacer(1, 0.2*inch))
        story.append(KeepTogether(block[:7])); story.extend(block[7:])

    doc.build(story, canvasmaker=NumCanvas)


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n RingCentral ACE Engine → http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
